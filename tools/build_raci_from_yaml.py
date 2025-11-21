#!/usr/bin/env python3
"""
build_raci_from_yaml.py

Generate an Excel RACI matrix from a YAML definition.

Usage:
    python build_raci_from_yaml.py --input raci_input.yaml --output raci_matrix.xlsx

Expected YAML structure:

    roles:
      - IR Lead
      - IT Team
      - Senior Management
      - DPO
      - MSP
      - Communications Lead

    tasks:
      - name: Incident Identification
        description: "Detect and recognise potential security incidents."
        assignments:
          IR Lead: R
          IT Team: C
          Senior Management: I
          DPO: I
          MSP: I
          Communications Lead: I

      - name: Incident Logging
        assignments:
          IR Lead: A
          IT Team: R

RACI codes:
    R = Responsible
    A = Accountable
    C = Consulted
    I = Informed

Any missing role for a task will be left blank in the matrix.
"""

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List

import yaml  # Requires PyYAML
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build RACI matrix Excel from YAML definition.")
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        help="Path to YAML file defining roles and tasks.",
    )
    parser.add_argument(
        "--output",
        "-o",
        required=False,
        default="RACI_Matrix_Generated.xlsx",
        help="Output Excel file path (default: RACI_Matrix_Generated.xlsx).",
    )
    return parser.parse_args()


def load_yaml(path: str) -> Dict[str, Any]:
    p = Path(path)
    if not p.exists():
        print(f"Error: input YAML not found: {path}", file=sys.stderr)
        sys.exit(1)
    try:
        with p.open("r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
    except yaml.YAMLError as e:
        print(f"Error parsing YAML file: {e}", file=sys.stderr)
        sys.exit(1)

    if "roles" not in data or "tasks" not in data:
        print("Error: YAML must contain 'roles' and 'tasks' keys.", file=sys.stderr)
        sys.exit(1)

    if not isinstance(data["roles"], list) or not isinstance(data["tasks"], list):
        print("Error: 'roles' must be a list and 'tasks' must be a list.", file=sys.stderr)
        sys.exit(1)

    return data


def build_raci_excel(data: Dict[str, Any], output_path: str) -> None:
    roles: List[str] = [str(r) for r in data.get("roles", [])]
    tasks: List[Dict[str, Any]] = data.get("tasks", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "RACI Matrix"

    # Header row: Task / Activity + roles
    header = ["Task / Activity"] + roles
    ws.append(header)

    # Style header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Color presets for RACI codes
    fill_R = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    fill_A = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    fill_C = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    fill_I = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Blue

    # Populate tasks
    for task in tasks:
        name = str(task.get("name", ""))
        assignments: Dict[str, str] = task.get("assignments", {}) or {}
        row_values = [name]
        for role in roles:
            value = assignments.get(role, "")
            row_values.append(value)
        ws.append(row_values)

    # Apply styling to RACI cells
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            value = (cell.value or "").strip().upper()
            if value == "R":
                cell.fill = fill_R
                cell.font = Font(bold=True)
            elif value == "A":
                cell.fill = fill_A
                cell.font = Font(bold=True)
            elif value == "C":
                cell.fill = fill_C
            elif value == "I":
                cell.fill = fill_I
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value))
            except Exception:
                cell_len = 0
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)


def main() -> None:
    args = parse_args()
    data = load_yaml(args.input)
    build_raci_excel(data, args.output)
    print(f"RACI matrix generated: {args.output}")


if __name__ == "__main__":
    main()
